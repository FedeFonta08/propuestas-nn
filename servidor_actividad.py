"""
NN ACTIVIDAD SERVER — servidor_actividad.py
=============================================
Servidor local que recibe registros de llamadas desde el Cockpit
y los escribe en la pestaña "Actividad Agentes" del Excel SISTEMA MAESTRO.

USO:
  python servidor_actividad.py

El servidor queda escuchando en http://localhost:5151
Guarda y libera el Excel en cada escritura (no lo bloquea).

REQUISITOS:
  pip install openpyxl
"""

import json
import urllib.parse
from http.server import HTTPServer, BaseHTTPRequestHandler
from datetime import datetime
import openpyxl
import os

# ── CONFIGURACIÓN ─────────────────────────────────────────────
EXCEL_PATH = r'D:\Users\ffont\Downloads\NN_SISTEMA_MAESTRO_2026_v2.xlsx'

PORT = 5151
TAB_ACTIVIDAD = 'Actividad Agentes'

# Fila en el Excel donde empieza cada agente (1-indexed para openpyxl)
# FEDE:    fila 7 → Llamadas | fila 8 → Contactos | fila 9 → Reuniones
# ROSIANE: fila 16 → Llamadas | fila 17 → Contactos | fila 18 → Reuniones
FILAS_AGENTE = {
    'FEDE':    {'llamadas': 7,  'contactos': 8,  'reuniones': 9,  'presupuestos': 10},
    'ROSIANE': {'llamadas': 16, 'contactos': 17, 'reuniones': 18, 'presupuestos': 19},
}

# Columna por día: B=2(LUN), C=3(MAR), D=4(MIE), E=5(JUE), F=6(VIE)
COL_DIA   = {'LUN': 2, 'MAR': 3, 'MIE': 4, 'JUE': 5, 'VIE': 6}
COL_TOTAL = 7  # Columna G = TOTAL

# ── LÓGICA EXCEL ─────────────────────────────────────────────
def registrar_actividad(agente, resultado, dia):
    agente = agente.upper()
    dia    = dia.upper()

    if agente not in FILAS_AGENTE:
        return {'ok': False, 'error': f'Agente desconocido: {agente}'}

    col_dia = COL_DIA.get(dia)
    if not col_dia:
        return {'ok': True, 'msg': f'Dia {dia} no laboral — sin registro'}

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        return {'ok': False, 'error': f'Excel no encontrado en: {EXCEL_PATH}'}

    if TAB_ACTIVIDAD not in wb.sheetnames:
        wb.close()
        return {'ok': False, 'error': f'Pestana "{TAB_ACTIVIDAD}" no encontrada'}

    ws    = wb[TAB_ACTIVIDAD]
    filas = FILAS_AGENTE[agente]

    def sumar(fila, col):
        cell      = ws.cell(row=fila, column=col)
        val       = cell.value
        cell.value = (int(val) if isinstance(val, (int, float)) else 0) + 1

    def recalc_total(fila):
        total = 0
        for c in COL_DIA.values():
            v = ws.cell(row=fila, column=c).value
            total += int(v) if isinstance(v, (int, float)) else 0
        ws.cell(row=fila, column=COL_TOTAL).value = total

    # Siempre: +1 llamada
    sumar(filas['llamadas'], col_dia)
    recalc_total(filas['llamadas'])

    # Si contactado, cita o presupuesto: +1 contacto
    if resultado in ('contactado', 'cita_fijada', 'presupuesto_enviado'):
        sumar(filas['contactos'], col_dia)
        recalc_total(filas['contactos'])

    # Si cita: +1 reunión
    if resultado == 'cita_fijada':
        sumar(filas['reuniones'], col_dia)
        recalc_total(filas['reuniones'])

    # Si presupuesto enviado: +1 presupuesto
    if resultado == 'presupuesto_enviado':
        sumar(filas['presupuestos'], col_dia)
        recalc_total(filas['presupuestos'])

    wb.save(EXCEL_PATH)
    wb.close()

    hora = datetime.now().strftime('%H:%M:%S')
    print(f'  [{hora}] OK {agente} | {resultado} | {dia} -> Excel actualizado')
    return {'ok': True, 'msg': f'{agente} {resultado} el {dia} registrado'}

def get_marcador_data():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        if TAB_ACTIVIDAD not in wb.sheetnames:
            return {'ok': False, 'error': 'Pestana no encontrada'}
        
        ws = wb[TAB_ACTIVIDAD]
        data = {'ok': True, 'FEDE': {}, 'ROSIANE': {}}
        
        for ag, filas in FILAS_AGENTE.items():
            data[ag] = {
                'llamadas':     int(ws.cell(row=filas['llamadas'], column=COL_TOTAL).value or 0),
                'contactos':    int(ws.cell(row=filas['contactos'], column=COL_TOTAL).value or 0),
                'reuniones':    int(ws.cell(row=filas['reuniones'], column=COL_TOTAL).value or 0),
                'presupuestos': int(ws.cell(row=filas['presupuestos'], column=COL_TOTAL).value or 0)
            }
        wb.close()
        return data
    except Exception as e:
        return {'ok': False, 'error': str(e)}

def get_pvm_data():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        # Intentar buscar en Dashboard o PVM
        sheet_name = 'Dashboard' if 'Dashboard' in wb.sheetnames else ('PVM' if 'PVM' in wb.sheetnames else None)
        if not sheet_name:
            return {'ok': False, 'error': 'Pestana PVM/Dashboard no encontrada'}
        
        ws = wb[sheet_name]
        # Asumimos que el PVM actual está en una celda identificable. 
        # Si no la conocemos, buscamos el texto "TOTAL PVM" o similar.
        current_pvm = 0
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'TOTAL' in cell.value.upper() and 'PVM' in cell.value.upper():
                    # El valor suele estar a la derecha
                    current_pvm = ws.cell(row=cell.row, column=cell.column + 1).value
                    break
            if current_pvm: break
        
        # Si no se encuentra por búsqueda, probar celda fija común (ej: B2 en PVM)
        if not current_pvm and sheet_name == 'PVM':
            current_pvm = ws['B2'].value

        wb.close()
        return {'ok': True, 'pvm': int(current_pvm or 0)}
    except Exception as e:
        return {'ok': False, 'error': str(e)}

# ── SERVIDOR HTTP ─────────────────────────────────────────────
class ActividadHandler(BaseHTTPRequestHandler):

    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()

        try:
            parsed = urllib.parse.urlparse(self.path)
            params = urllib.parse.parse_qs(parsed.query)
            raw    = params.get('payload', ['{}'])[0]
            data   = json.loads(urllib.parse.unquote(raw))

            if data.get('accion') == 'registrar_actividad':
                result = registrar_actividad(
                    agente    = data.get('agente', ''),
                    resultado = data.get('resultado', ''),
                    dia       = data.get('dia', '')
                )
            elif data.get('accion') == 'get_marcador':
                result = get_marcador_data()
            elif data.get('accion') == 'get_pvm':
                result = get_pvm_data()
            else:
                result = {'ok': False, 'error': 'Accion desconocida'}

        except Exception as e:
            result = {'ok': False, 'error': str(e)}

        self.wfile.write(json.dumps(result).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.end_headers()

    def log_message(self, format, *args):
        pass  # Silenciar logs HTTP estandar


# ── MAIN ──────────────────────────────────────────────────────
if __name__ == '__main__':
    print('\n' + '='*52)
    print('  NN ACTIVIDAD SERVER — Sprint 5')
    print('='*52)
    print(f'  Puerto : http://localhost:{PORT}')
    print(f'  Excel  : {EXCEL_PATH}')
    print(f'  Pestana: {TAB_ACTIVIDAD}')
    print('='*52)

    if not os.path.exists(EXCEL_PATH):
        print(f'\n  AVISO: Excel no encontrado. Edita EXCEL_PATH en este script.\n')
    else:
        print(f'  Excel encontrado. Listo.\n')

    print('  Servidor arrancado. Ctrl+C para detener.\n')

    server = HTTPServer(('localhost', PORT), ActividadHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\n  Servidor detenido.')
        server.server_close()
